from __future__ import annotations

import pytest
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from pif_research_platform.api import create_app
from pif_research_platform.models import (
    AnalysisPack,
    AnalysisDataPoint,
    CheckpointDecision,
    CheckpointId,
    CheckpointSubmission,
    OutputFormat,
    RunCreateRequest,
    RunStageId,
    SerializedRunState,
)
from pif_research_platform.service import build_default_service
from pif_research_platform.utils import word_count


GENERIC_TOPIC = "Assess EV charging readiness in Karnataka using infrastructure, demand, and financing indicators"
ENVIRONMENT_TOPIC = "Access air pollution report in Delhi, India"


@pytest.fixture(autouse=True)
def _force_template_llm(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("PIF_LOCAL_LLM_PROVIDER", "template")
    monkeypatch.setenv("PIF_LOCAL_LLM_FALLBACK", "template")
    monkeypatch.setenv("PIF_SEARCH_PROVIDER", "offline")
    monkeypatch.setenv("PIF_SEARCH_FALLBACK", "offline")


def test_end_to_end_run_generates_required_outputs(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=GENERIC_TOPIC,
            output_format=OutputFormat.MARKDOWN,
            target_word_count=1800,
        )
    )
    run_id = summary.run_id
    assert summary.active_checkpoint == CheckpointId.CONFIG.value

    for checkpoint in [CheckpointId.CONFIG, CheckpointId.PLAN, CheckpointId.DATA, CheckpointId.ANALYSIS]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )
        if checkpoint == CheckpointId.CONFIG:
            state = service.repo.load_state(run_id)
            assert state.indicator_plan
            assert all("example.org" not in item.primary_source_url for item in state.indicator_plan)
            assert all(item.primary_source_url.startswith("https://") for item in state.indicator_plan)
            assert all(item.verification_source_url.startswith("https://") for item in state.indicator_plan)

    assert summary.active_checkpoint == CheckpointId.REVIEW.value
    assert summary.review_cycles == 1

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Incorporate the recommended fixes.",
        ),
    )

    assert summary.status.value == "completed"
    assert summary.active_checkpoint is None
    assert summary.latest_score is not None and summary.latest_score >= 7.5
    assert summary.review_cycles == 2

    report_path = Path(summary.artifact_paths["report"])
    workbook_path = Path(summary.artifact_paths["indicator_workbook"])
    final_scores_path = Path(summary.artifact_paths["final_scores"])
    assert report_path.exists()
    assert workbook_path.exists()
    assert final_scores_path.exists()

    report_text = report_path.read_text(encoding="utf-8")
    assert 1620 <= word_count(report_text) <= 1980
    assert "Maharashtra" not in report_text
    assert "Karnataka" in report_text
    workbook = load_workbook(workbook_path)
    assert {"indicator_table", "metadata", "analysis_output"}.issubset(set(workbook.sheetnames))


def test_checkpoint_revisions_restart_correct_stage(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(RunCreateRequest(topic=GENERIC_TOPIC))
    run_id = summary.run_id

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.CONFIG,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.PLAN.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.PLAN,
            decision=CheckpointDecision.REVISE,
            feedback="Clarify the analytical method.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.PLAN.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.PLAN,
            decision=CheckpointDecision.APPROVE,
            feedback="Looks good now.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.DATA.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.DATA,
            decision=CheckpointDecision.REVISE,
            feedback="Re-run the data verification summary.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.DATA.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.DATA,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.ANALYSIS.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.ANALYSIS,
            decision=CheckpointDecision.REVISE,
            feedback="Clarify the interval interpretation.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.ANALYSIS.value


def test_macro_indicator_plan_uses_real_official_urls(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(RunCreateRequest(topic="Analysis latest gdsp of delhi"))
    service.submit_checkpoint(
        summary.run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.CONFIG,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed",
        ),
    )
    state = service.repo.load_state(summary.run_id)
    urls = {item.indicator_id: item.primary_source_url for item in state.indicator_plan}
    verify_urls = {item.indicator_id: item.verification_source_url for item in state.indicator_plan}

    assert state.topic_profile is not None
    assert state.topic_profile.geography == "Delhi"
    assert urls["TAX"].startswith("https://")
    assert urls["POWER"].startswith("https://")
    assert urls["REG"].startswith("https://")
    assert verify_urls["CREDIT"].startswith("https://")
    assert all("example.org" not in url for url in list(urls.values()) + list(verify_urls.values()))


def test_environment_topic_uses_environment_profile_and_specific_recommendations(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=ENVIRONMENT_TOPIC,
            output_format=OutputFormat.MARKDOWN,
            target_word_count=1800,
        )
    )
    run_id = summary.run_id

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.CONFIG,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed",
        ),
    )
    state = service.repo.load_state(run_id)
    assert state.topic_profile is not None
    assert state.topic_profile.domain == "environment"
    assert {item.indicator_id for item in state.indicator_plan} >= {"AQI", "PM25", "COMPLIANCE"}

    for checkpoint in [CheckpointId.PLAN, CheckpointId.DATA, CheckpointId.ANALYSIS]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Apply any final fixes.",
        ),
    )

    report_text = Path(summary.artifact_paths["report"]).read_text(encoding="utf-8")
    report_lower = report_text.lower()
    assert "environmental risk score" in report_lower
    assert "delhi pollution control committee" in report_lower
    assert "economic momentum estimate" not in report_lower
    assert "tax collections growth" not in report_lower


def test_long_report_expansion_does_not_repeat_identical_watchlist_text(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=ENVIRONMENT_TOPIC,
            output_format=OutputFormat.MARKDOWN,
            target_word_count=3000,
        )
    )
    run_id = summary.run_id

    for checkpoint in [CheckpointId.CONFIG, CheckpointId.PLAN, CheckpointId.DATA, CheckpointId.ANALYSIS]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Apply final fixes if needed.",
        ),
    )
    report_text = Path(summary.artifact_paths["report"]).read_text(encoding="utf-8")
    assert 2700 <= word_count(report_text) <= 3300
    assert report_text.count("### Monitoring Watchlist") == 1
    assert report_text.count("The monitoring watchlist for") == 1


def test_planner_uses_llm_defined_dynamic_domain_and_indicator_basket(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = build_default_service(tmp_path)

    def fake_complete_json(self, *, system_prompt: str, user_prompt: str, fallback):
        if "design topic-specific research profiles" in system_prompt:
            return {
                "domain": "judicial_delivery",
                "geography": "Uttar Pradesh",
                "research_goal": "Assess judicial case backlog reduction capacity, operational bottlenecks, and implementation priorities for the topic.",
                "headline_metric_label": "Justice delivery performance score",
                "headline_unit": "index",
                "comparator_label": "Previous court administration baseline",
                "dimension_label": "Court-system driver attribution",
                "keywords": ["judicial backlog", "pendency", "case clearance", "court staffing"],
                "indicator_dimensions": ["Pendency", "Clearance", "Capacity", "Digital", "Access"],
                "source_themes": ["case management reform", "court administration", "judicial delays", "service delivery"],
            }
        if "design indicator baskets" in system_prompt:
            return {
                "indicators": [
                    {
                        "indicator_id": "PENDENCY",
                        "label": "Case pendency trend",
                        "frequency": "quarterly",
                        "unit": "index",
                        "sector_bucket": "Pendency",
                        "rationale": "Tracks whether unresolved case stock is improving or worsening.",
                    },
                    {
                        "indicator_id": "CLEARANCE",
                        "label": "Case clearance rate",
                        "frequency": "quarterly",
                        "unit": "% share",
                        "sector_bucket": "Clearance",
                        "rationale": "Measures how much incoming caseload is being disposed in time.",
                    },
                    {
                        "indicator_id": "VACANCY",
                        "label": "Judge vacancy pressure",
                        "frequency": "quarterly",
                        "unit": "% share",
                        "sector_bucket": "Capacity",
                        "rationale": "Captures staffing gaps affecting disposal capacity.",
                    },
                    {
                        "indicator_id": "ECOURTS",
                        "label": "Digital filing adoption",
                        "frequency": "quarterly",
                        "unit": "% share",
                        "sector_bucket": "Digital",
                        "rationale": "Signals operational modernization of court workflows.",
                    },
                    {
                        "indicator_id": "LEGALAID",
                        "label": "Legal aid access trend",
                        "frequency": "quarterly",
                        "unit": "index",
                        "sector_bucket": "Access",
                        "rationale": "Tracks whether service access is improving for vulnerable users.",
                    },
                ]
            }
        return fallback

    monkeypatch.setattr(type(service.agent_suite.local_llm), "complete_json", fake_complete_json)

    summary = service.create_run(RunCreateRequest(topic="Reduce judicial case backlog in Uttar Pradesh"))
    service.submit_checkpoint(
        summary.run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.CONFIG,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed",
        ),
    )
    state = service.repo.load_state(summary.run_id)

    assert state.topic_profile is not None
    assert state.topic_profile.domain == "judicial_delivery"
    assert state.topic_profile.headline_metric_label == "Justice delivery performance score"
    assert state.topic_profile.indicator_dimensions[:3] == ["Pendency", "Clearance", "Capacity"]
    assert {item.indicator_id for item in state.indicator_plan} >= {"PENDENCY", "CLEARANCE", "VACANCY"}
    assert all("signal" not in item.label.lower() for item in state.indicator_plan[:3])


def test_writer_validation_blocks_unverified_references(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    state = SerializedRunState(
        run_id="validation",
        topic="topic",
        output_format=OutputFormat.MARKDOWN,
        target_word_count=1800,
        verified_allowlist=["S1"],
        analysis_pack=AnalysisPack(
            target_period="2026-Q1",
            model_name="model",
            headline_metric_label="Transition readiness score",
            headline_unit="index",
            comparator_label="Previous readiness baseline",
            dimension_label="Capability attribution",
            point_estimate=7.2,
            confidence_low=6.8,
            confidence_high=7.6,
            last_official_estimate=6.9,
            sector_attribution={"Demand": 0.5},
            scenario_commentary="scenario",
            method_notes="notes",
            data_points=[
                AnalysisDataPoint(
                    datapoint_id="A1",
                    label="label",
                    value=1.0,
                    unit="% YoY",
                    explanation="ok",
                )
            ],
            analysis_table=[{"bucket": "Demand", "contribution": 0.5, "interpretation": "Positive"}],
        ),
    )

    try:
        service.agent_suite.validate_report_references("This cites [S2] and [A2].", state)
    except ValueError as exc:
        assert "unverified references" in str(exc)
    else:  # pragma: no cover
        raise AssertionError("Expected writer validation to fail")


def test_review_loop_stops_after_max_cycles(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=GENERIC_TOPIC,
            notes="[always-fail-review]",
        )
    )
    run_id = summary.run_id
    for checkpoint in [CheckpointId.CONFIG, CheckpointId.PLAN, CheckpointId.DATA, CheckpointId.ANALYSIS]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )
    assert summary.active_checkpoint == CheckpointId.REVIEW.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Try another revision pass.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.REVIEW.value

    summary = service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Final permitted revision pass.",
        ),
    )
    assert summary.status.value == "completed"
    assert summary.review_cycles == 3
    assert summary.active_checkpoint is None


def test_api_endpoints_match_service_contract(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    client = TestClient(create_app(service))
    response = client.post(
        "/runs",
        json={
            "topic": GENERIC_TOPIC,
            "output_format": "markdown",
            "target_word_count": 1800,
        },
    )
    assert response.status_code == 200
    payload = response.json()
    run_id = payload["run_id"]
    assert payload["active_checkpoint"] == "config"

    response = client.get(f"/runs/{run_id}")
    assert response.status_code == 200
    assert response.json()["run_id"] == run_id

    response = client.get(f"/runs/{run_id}/detail")
    assert response.status_code == 200
    assert response.json()["run_id"] == run_id


def test_restart_from_stage_truncates_later_outputs(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=GENERIC_TOPIC,
            output_format=OutputFormat.MARKDOWN,
            target_word_count=1800,
        )
    )
    run_id = summary.run_id

    for checkpoint in [CheckpointId.CONFIG, CheckpointId.PLAN, CheckpointId.DATA, CheckpointId.ANALYSIS]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )
    assert summary.active_checkpoint == CheckpointId.REVIEW.value

    summary = service.restart_from_stage(run_id, RunStageId.ANALYSIS)
    assert summary.active_checkpoint == CheckpointId.ANALYSIS.value
    assert summary.current_node == "analysis_checkpoint"

    state = service.get_run_detail(run_id)
    assert state.analysis_pack is not None
    assert state.report_versions == []
    assert state.qa_pack is None
    assert state.review_cycles == []
    assert "report" not in state.artifact_paths
    assert "final_scores" not in state.artifact_paths


def test_checkpoint_resume_survives_service_restart(tmp_path: Path) -> None:
    service = build_default_service(tmp_path)
    summary = service.create_run(
        RunCreateRequest(
            topic=GENERIC_TOPIC,
            output_format=OutputFormat.MARKDOWN,
            target_word_count=1800,
        )
    )
    run_id = summary.run_id

    for checkpoint in [CheckpointId.CONFIG, CheckpointId.PLAN, CheckpointId.DATA]:
        summary = service.submit_checkpoint(
            run_id,
            CheckpointSubmission(
                checkpoint_id=checkpoint,
                decision=CheckpointDecision.APPROVE,
                feedback="Proceed",
            ),
        )
    assert summary.active_checkpoint == CheckpointId.ANALYSIS.value

    restarted_service = build_default_service(tmp_path)
    summary = restarted_service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.ANALYSIS,
            decision=CheckpointDecision.APPROVE,
            feedback="Proceed to review.",
        ),
    )
    assert summary.active_checkpoint == CheckpointId.REVIEW.value
    assert summary.status.value == "waiting_for_checkpoint"

    restarted_again_service = build_default_service(tmp_path)
    summary = restarted_again_service.submit_checkpoint(
        run_id,
        CheckpointSubmission(
            checkpoint_id=CheckpointId.REVIEW,
            decision=CheckpointDecision.APPROVE,
            feedback="Apply the reviewer fixes.",
        ),
    )
    assert summary.status.value == "completed"
    assert summary.active_checkpoint is None
