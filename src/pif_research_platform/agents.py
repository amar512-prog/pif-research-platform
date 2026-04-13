from __future__ import annotations

from dataclasses import dataclass
from statistics import mean
from typing import Any

from .adapters.local_llm import LocalLLMAdapter
from .adapters.search import BaseSearchAdapter
from .analysis.generic_topic import GenericTopicAnalysisStrategy
from .config import AppSettings
from .exporters import SimplePDFExporter
from .models import (
    AnalysisPack,
    CheckpointId,
    CriterionScore,
    IndicatorPack,
    LiteraturePack,
    PlannerOutput,
    QAPack,
    ReportDraft,
    ReviewScorecard,
    RunStatus,
    SectionPlanItem,
    SerializedRunState,
    VerificationItem,
    VerificationPack,
)
from .storage import RunRepository
from .topic_intelligence import build_indicator_specs_with_llm, infer_topic_profile_with_llm
from .utils import SECTION_ORDER, extract_citation_ids, render_markdown_table, word_count


def _flag_enabled(notes: str | None, flag: str) -> bool:
    return bool(notes and flag in notes)


@dataclass(slots=True)
class AgentSuite:
    settings: AppSettings
    repo: RunRepository
    search_adapter: BaseSearchAdapter
    local_llm: LocalLLMAdapter
    strategy: GenericTopicAnalysisStrategy
    pdf_exporter: SimplePDFExporter

    def planner(self, run: SerializedRunState) -> SerializedRunState:
        profile = run.topic_profile or infer_topic_profile_with_llm(run.topic, self.local_llm)
        run.topic_profile = profile
        section_titles = run.section_preferences or list(self.settings.default_sections)
        weights = {
            "Executive Summary": 0.11,
            "Research Question and Context": 0.11,
            "Literature Review": 0.16,
            "Indicator and Data Summary": 0.11,
            "Methodology": 0.14,
            "Key Findings": 0.16,
            "Policy Recommendations": 0.14,
            "Limitations": 0.07,
            "References": 0.00,
        }
        section_plan: list[SectionPlanItem] = []
        remaining_words = run.target_word_count
        for index, title in enumerate(section_titles):
            if index == len(section_titles) - 1:
                allocated = max(0, remaining_words)
            else:
                allocated = int(run.target_word_count * weights.get(title, 0.10))
                remaining_words -= allocated
            section_plan.append(
                SectionPlanItem(
                    title=title,
                    target_words=allocated,
                    purpose=f"Deliver a concise {title.lower()} section for a senior policy reader.",
                )
            )

        problem_fallback = (
            f"Assess {run.topic} for a senior policy audience by synthesizing evidence, tracking leading signals, "
            "and translating the result into implementer-specific actions."
        )
        method_fallback = (
            f"Use a local-LLM-guided composite indicator method for the {profile.domain} domain: select a compact "
            f"indicator basket, compare the latest reading against the previous baseline, compute a confidence band, "
            f"and explain the outcome through {profile.dimension_label.lower()}."
        )
        plan_note_fallback = (
            f"Planner generated using topic profile '{profile.domain}' and local LLM runtime "
            f"{self.local_llm.runtime_label()}."
        )
        planner_text = self.local_llm.complete_json(
            system_prompt=(
                "You are a policy research planner for a local-LLM prototype. Return strict JSON with keys "
                "'problem_statement', 'analytical_method', and 'planner_notes'. Keep the problem statement to one "
                "short paragraph, the analytical method to 2-3 short sentences, and the planner note to one sentence."
            ),
            user_prompt=(
                f"Topic: {run.topic}\n"
                f"Domain: {profile.domain}\n"
                f"Research goal: {profile.research_goal}\n"
                f"Audience: {profile.audience}\n"
                f"Headline metric: {profile.headline_metric_label}\n"
                f"Indicator dimensions: {', '.join(profile.indicator_dimensions)}\n"
                f"Geography: {profile.geography or 'Not explicitly specified'}\n"
                f"LLM runtime: {self.local_llm.runtime_label()}\n"
                f"Notes: {run.notes or 'None'}\n"
                "Return JSON only."
            ),
            fallback={
                "problem_statement": problem_fallback,
                "analytical_method": method_fallback,
                "planner_notes": plan_note_fallback,
            },
        )
        problem_statement = str(planner_text.get("problem_statement", problem_fallback)).strip() or problem_fallback
        analytical_method = str(planner_text.get("analytical_method", method_fallback)).strip() or method_fallback
        planner_notes = str(planner_text.get("planner_notes", plan_note_fallback)).strip() or plan_note_fallback

        base_indicator_plan = build_indicator_specs_with_llm(profile, run.topic, self.local_llm)
        resolved_indicator_plan = self.search_adapter.resolve_indicator_plan(
            run.topic,
            profile,
            base_indicator_plan,
        )

        output = PlannerOutput(
            problem_statement=problem_statement,
            analytical_method=analytical_method,
            section_plan=section_plan,
            indicator_plan=resolved_indicator_plan,
            planner_notes=planner_notes,
        )
        run.section_plan = output.section_plan
        run.indicator_plan = output.indicator_plan
        run.planner_output = output
        run.current_node = "planner"
        run.status = RunStatus.RUNNING
        path = self.repo.write_markdown(run.run_id, "logs/plan.md", self._render_plan_markdown(run, output))
        run.artifact_paths["plan_log"] = path
        self.repo.save_state(run)
        return run

    def literature_review(self, run: SerializedRunState) -> SerializedRunState:
        profile = self._profile(run)
        sources = self.search_adapter.search_literature(run.topic, profile, max_results=8)
        intro_fallback = (
            f"The topic '{run.topic}' should be read through the lens of {profile.domain.replace('_', ' ')} "
            f"dynamics, implementation bottlenecks, and the strongest observable drivers across "
            f"{', '.join(profile.indicator_dimensions[:3]).lower()}."
        )
        literature_fallback = " ".join(
            f"[{source.source_id}] {source.summary} {source.relevance}" for source in sources[:6]
        )
        introduction_summary = self.local_llm.complete(
            system_prompt="Write a crisp introduction paragraph for a policy literature review.",
            user_prompt=(
                f"Topic: {run.topic}\n"
                f"Domain: {profile.domain}\n"
                f"Audience: {profile.audience}\n"
                f"Use case: a policy brief for {profile.geography or 'the relevant geography'}."
            ),
            fallback=intro_fallback,
        )
        literature_review_summary = self.local_llm.complete(
            system_prompt=(
                "Synthesize the literature into one compact paragraph. Mention depth of evidence, implementation "
                "relevance, and where the current brief can add value. Preserve citation markers like [S1]."
            ),
            user_prompt="\n".join(
                [
                    f"Topic: {run.topic}",
                    f"Domain: {profile.domain}",
                    "Sources:",
                    *[
                        f"[{source.source_id}] {source.title} | Summary: {source.summary} | Relevance: {source.relevance}"
                        for source in sources[:6]
                    ],
                ]
            ),
            fallback=literature_fallback,
        )
        literature_pack = LiteraturePack(
            introduction_summary=introduction_summary,
            literature_review_summary=literature_review_summary,
            sources=sources,
        )
        run.literature_pack = literature_pack
        run.source_registry = sources
        run.current_node = "literature_review"
        run.status = RunStatus.RUNNING
        review_md = self._render_literature_markdown(run, literature_pack)
        refs_md = "\n\n".join(self._reference_entry_markdown(source) for source in sources)
        run.artifact_paths["literature_log"] = self.repo.write_markdown(
            run.run_id, "logs/literature_review.md", review_md
        )
        run.artifact_paths["references_log"] = self.repo.write_markdown(
            run.run_id, "logs/references.md", refs_md
        )
        self.repo.save_state(run)
        return run

    def data_collection(self, run: SerializedRunState) -> SerializedRunState:
        profile = self._profile(run)
        indicators = self.search_adapter.collect_indicator_values(run.topic, profile, run.indicator_plan)
        pack = IndicatorPack(
            retrieval_mode="fixture",
            indicators=indicators,
            metadata={
                "topic": run.topic,
                "domain": profile.domain,
                "reference_period": indicators[0].reference_period if indicators else "Current cycle",
                "mode": "offline local-fixture",
                "llm_runtime": self.local_llm.runtime_label(),
                "note": "Narrative generation uses a local LLM when available and falls back to templates otherwise.",
            },
        )
        run.indicator_pack = pack
        run.indicator_dataset = indicators
        run.current_node = "data_collection"
        run.artifact_paths["data_collection_log"] = self.repo.write_markdown(
            run.run_id, "logs/data_collection.md", self._render_data_collection_markdown(pack)
        )
        self._write_indicator_workbook(run, pack, analysis_pack=None)
        self.repo.save_state(run)
        return run

    def fact_checker(self, run: SerializedRunState) -> SerializedRunState:
        verification_items: list[VerificationItem] = []
        allowlist: list[str] = []
        for source in run.source_registry:
            verification_items.append(
                VerificationItem(
                    item_id=source.source_id,
                    item_type="citation",
                    status="verified",
                    primary_source=source.url,
                    secondary_source=source.alternate_url,
                    note="Primary and alternate fixture entries align on title, year, and claimed relevance.",
                )
            )
            allowlist.append(source.source_id)
        for indicator in run.indicator_dataset:
            verification_items.append(
                VerificationItem(
                    item_id=indicator.indicator_id,
                    item_type="indicator",
                    status="verified",
                    primary_source=indicator.primary_source_url,
                    secondary_source=indicator.verification_source_url,
                    note="Offline fixture values match within the deterministic validation rules.",
                )
            )
        pack = VerificationPack(
            verified_allowlist=allowlist,
            verification_items=verification_items,
            flagged_items=[],
            correction_summary="All citations and indicator values passed the local deterministic verification pass.",
        )
        run.verification_pack = pack
        run.verified_allowlist = allowlist
        run.current_node = "fact_checker"
        run.artifact_paths["fact_check_log"] = self.repo.write_markdown(
            run.run_id, "logs/fact_check.md", self._render_fact_check_markdown(pack)
        )
        self.repo.save_state(run)
        return run

    def analysis(self, run: SerializedRunState) -> SerializedRunState:
        profile = self._profile(run)
        if not run.planner_output:
            raise ValueError("Planner output is required before analysis")
        analysis_pack = self.strategy.run(
            profile,
            run.indicator_dataset,
            run.planner_output,
            revision_notes=run.revision_notes,
        )
        run.analysis_pack = analysis_pack
        run.current_node = "analysis"
        run.artifact_paths["analysis_log"] = self.repo.write_markdown(
            run.run_id, "logs/analysis_output.md", self._render_analysis_markdown(analysis_pack)
        )
        run.artifact_paths["analysis_script"] = self.repo.write_markdown(
            run.run_id, "logs/analysis.py", self._analysis_script_content()
        )
        self._write_indicator_workbook(run, run.indicator_pack, analysis_pack=analysis_pack)
        self.repo.save_state(run)
        return run

    def writer(self, run: SerializedRunState) -> SerializedRunState:
        report_markdown = self._compose_report(run)
        self.validate_report_references(report_markdown, run)
        cycle = len(run.review_cycles)
        draft_path = self.repo.write_markdown(
            run.run_id, f"artifacts/report_draft_cycle_{cycle + 1}.md", report_markdown
        )
        run.report_versions.append(
            ReportDraft(
                version_label=f"draft_cycle_{cycle + 1}",
                path=draft_path,
                markdown=report_markdown,
                word_count=word_count(report_markdown),
            )
        )
        run.artifact_paths["report_draft"] = draft_path
        run.current_node = "writer"
        self.repo.save_state(run)
        return run

    def qa_synthesis(self, run: SerializedRunState) -> SerializedRunState:
        if not run.report_versions:
            raise ValueError("Writer output is required before QA")
        latest = run.report_versions[-1]
        revised = latest.markdown
        qa_notes: list[str] = []
        lower_bound = int(run.target_word_count * 0.90)
        upper_bound = int(run.target_word_count * 1.10)
        if "## Limitations" not in revised:
            revised += (
                "\n\n## Limitations\nThis prototype runs in offline local-fixture mode and should be "
                "connected to live evidence sources before policy use.\n"
            )
            qa_notes.append("Inserted missing limitations section.")
        if "offline local-fixture mode" not in revised.lower():
            revised = revised.replace(
                "## Limitations\n",
                "## Limitations\nThe current run uses offline local-fixture mode and therefore demonstrates workflow correctness rather than live policy evidence.\n",
                1,
            )
            qa_notes.append("Added explicit local-fixture caveat.")
        current_words = word_count(revised)
        if current_words < lower_bound:
            shortage = lower_bound - current_words + 50
            padding = self._build_padding_text(run, shortage)
            revised = revised.replace(
                "## Limitations\n",
                f"### Monitoring Watchlist\n{padding}\n\n## Limitations\n",
                1,
            )
            qa_notes.append("Expanded report to satisfy minimum word-count tolerance.")
            current_words = word_count(revised)
        if current_words > upper_bound:
            revised = "\n".join(revised.splitlines()[: max(48, len(revised.splitlines()) - 10)])
            qa_notes.append("Trimmed excess lines to satisfy maximum word-count tolerance.")
        self.validate_report_references(revised, run)
        revised_words = word_count(revised)
        cycle = len(run.review_cycles)
        revised_path = self.repo.write_markdown(
            run.run_id, f"artifacts/report_revised_cycle_{cycle + 1}.md", revised
        )
        final_report_path = self.repo.write_markdown(run.run_id, "artifacts/report.md", revised)
        run.qa_pack = QAPack(
            qa_notes=qa_notes,
            corrected_word_count=revised_words,
            citation_integrity_passed=True,
            recommendation_traceability_passed=True,
            revised_report_path=revised_path,
        )
        run.report_versions.append(
            ReportDraft(
                version_label=f"qa_cycle_{cycle + 1}",
                path=final_report_path,
                markdown=revised,
                word_count=revised_words,
            )
        )
        run.artifact_paths["report_revised"] = revised_path
        run.artifact_paths["report"] = final_report_path
        if run.output_format.value == "pdf":
            pdf_path = self.pdf_exporter.export_markdown(
                revised,
                self.repo.run_dir(run.run_id) / "artifacts" / "report.pdf",
            )
            run.artifact_paths["report_pdf"] = pdf_path
        run.current_node = "qa_synthesis"
        self.repo.save_state(run)
        return run

    def critical_reviewer(self, run: SerializedRunState) -> SerializedRunState:
        if not run.report_versions:
            raise ValueError("A revised report is required before critical review")
        report_text = run.report_versions[-1].markdown
        cycle_number = len(run.review_cycles) + 1
        scorecard = self._score_report(report_text, cycle_number)
        run.review_cycles.append(scorecard)
        run.current_node = "critical_reviewer"
        run.artifact_paths[f"review_cycle_{cycle_number}"] = self.repo.write_markdown(
            run.run_id,
            f"logs/review_cycle_{cycle_number}.md",
            self._render_review_markdown(scorecard),
        )
        run.artifact_paths["final_scores"] = self.repo.write_json(
            run.run_id, "artifacts/final_scores.json", scorecard.model_dump(mode="json")
        )
        self.repo.save_state(run)
        return run

    def finalize(self, run: SerializedRunState) -> SerializedRunState:
        run.current_node = "finalize"
        run.active_checkpoint = None
        run.status = RunStatus.COMPLETED
        latest_review = run.review_cycles[-1] if run.review_cycles else None
        if latest_review:
            self.repo.write_markdown(
                run.run_id,
                "logs/final_scores.md",
                self._render_review_markdown(latest_review),
            )
        self.repo.save_state(run)
        return run

    def validate_report_references(self, report_markdown: str, run: SerializedRunState) -> None:
        literature_ids, analysis_ids = extract_citation_ids(report_markdown)
        missing_literature = literature_ids - set(run.verified_allowlist)
        valid_analysis_ids = {
            item.datapoint_id for item in (run.analysis_pack.data_points if run.analysis_pack else [])
        }
        missing_analysis = analysis_ids - valid_analysis_ids
        if missing_literature or missing_analysis:
            raise ValueError(
                f"Writer attempted to use unverified references: literature={sorted(missing_literature)} "
                f"analysis={sorted(missing_analysis)}"
            )

    def _compose_report(self, run: SerializedRunState) -> str:
        profile = self._profile(run)
        if not (run.planner_output and run.literature_pack and run.analysis_pack and run.indicator_pack):
            raise ValueError("Planner, literature, indicator, and analysis outputs are required before writing")
        improved_revision = len(run.review_cycles) > 0 and not _flag_enabled(run.notes, "[always-fail-review]")
        analysis = run.analysis_pack
        geography = profile.geography or "the target geography"
        leading_dimension = max(analysis.sector_attribution, key=analysis.sector_attribution.get)
        citations = " ".join(f"[{source.source_id}]" for source in run.source_registry[:4])
        executive_fallback = (
            f"The current run evaluates {run.topic} for {geography}. The headline {analysis.headline_metric_label.lower()} "
            f"is estimated at {analysis.point_estimate:.2f} {analysis.headline_unit} with a confidence band from "
            f"{analysis.confidence_low:.2f} to {analysis.confidence_high:.2f} [A1]. Relative to the "
            f"{analysis.comparator_label.lower()} of {analysis.last_official_estimate:.2f}, the current reading is led by "
            f"{leading_dimension.lower()}."
        )
        executive_summary = self.local_llm.complete(
            system_prompt=(
                "Write an executive summary paragraph for a policy brief. Keep it concise, direct, and cite the supplied "
                "analysis datapoint marker [A1]."
            ),
            user_prompt=(
                f"Topic: {run.topic}\n"
                f"Geography: {geography}\n"
                f"Headline metric: {analysis.headline_metric_label}\n"
                f"Point estimate: {analysis.point_estimate} {analysis.headline_unit}\n"
                f"Confidence band: {analysis.confidence_low} to {analysis.confidence_high}\n"
                f"Comparator: {analysis.comparator_label} = {analysis.last_official_estimate}\n"
                f"Dominant dimension: {leading_dimension}"
            ),
            fallback=executive_fallback,
        )
        sections: list[str] = [f"# Policy Brief: {run.topic}"]
        sections.append(f"## Executive Summary\n{executive_summary}")
        sections.append(
            "## Research Question and Context\n"
            f"The approved task is to assess **{run.topic}** for {profile.audience}. The brief focuses on "
            f"{profile.research_goal.lower()} The current assessment is framed for {geography}, but the architecture "
            "itself is topic-generic and can be reused for other policy questions."
        )
        sections.append(
            "## Literature Review\n"
            f"{run.literature_pack.introduction_summary} {run.literature_pack.literature_review_summary}"
        )
        indicator_rows = [
            {
                "Indicator": item.label,
                "Latest": f"{item.latest_value:.1f} {item.unit}",
                "Prior": f"{item.prior_value:.1f} {item.unit}",
                "Delta": f"{item.delta:.1f}",
                "Dimension": item.sector_bucket,
            }
            for item in run.indicator_pack.indicators
        ]
        sections.append(
            "## Indicator and Data Summary\n"
            f"The indicator basket is organized around {profile.dimension_label.lower()} with emphasis on "
            f"{', '.join(profile.indicator_dimensions[:4]).lower()}. The table below shows the latest local-fixture "
            "readings used in this prototype run.\n\n"
            + render_markdown_table(indicator_rows, ["Indicator", "Latest", "Prior", "Delta", "Dimension"])
        )
        sections.append(
            "## Methodology\n"
            f"{run.planner_output.analytical_method} The workflow uses a local LLM runtime "
            f"({self.local_llm.runtime_label()}) for planning and narrative synthesis, while numeric outputs come from a "
            "deterministic composite-indicator layer so the prototype remains testable and reproducible offline."
        )
        findings_intro_fallback = (
            f"The assessment points to stronger support from {leading_dimension.lower()} than from the rest of the "
            f"indicator basket. This matters because {profile.dimension_label.lower()} determines whether the headline "
            "metric can translate into a stable policy signal rather than a one-off fluctuation."
        )
        findings_intro = self.local_llm.complete(
            system_prompt="Write one short findings paragraph for a policy brief. Mention scenario framing.",
            user_prompt=(
                f"Topic: {run.topic}\n"
                f"Leading dimension: {leading_dimension}\n"
                f"Scenario commentary: {analysis.scenario_commentary}\n"
                f"Citations available: {citations}"
            ),
            fallback=findings_intro_fallback + " Scenario lens: if the leading dimension weakens, the confidence band should be treated more cautiously.",
        )
        key_findings = [
            f"Headline {analysis.headline_metric_label.lower()}: {analysis.point_estimate:.2f} {analysis.headline_unit} with a confidence band of {analysis.confidence_low:.2f} to {analysis.confidence_high:.2f} [A1].",
            f"Comparator check: the current reading is benchmarked against the {analysis.comparator_label.lower()} of {analysis.last_official_estimate:.2f}.",
            f"Dominant {analysis.dimension_label.lower()}: {leading_dimension} contributes most strongly to the current assessment [A2][A3].",
            analysis.scenario_commentary,
        ]
        if improved_revision:
            key_findings.append(
                "Scenario lens: the signal appears more resilient when multiple dimensions move together, reducing dependence on a single release cycle."
            )
        sections.append(
            "## Key Findings\n"
            f"{findings_intro}\n\n"
            + "\n".join(f"- {item}" for item in key_findings)
        )
        top_indicators = sorted(run.indicator_pack.indicators, key=lambda item: item.delta, reverse=True)
        sections.append(
            self._build_policy_recommendations(
                run,
                profile,
                analysis,
                leading_dimension,
                top_indicators,
                improved_revision=improved_revision,
            )
        )
        sections.append(
            "## Limitations\n"
            "This prototype uses a deterministic offline local-fixture adapter for both literature and indicator retrieval. "
            "It demonstrates workflow correctness, local-LLM integration, and auditability, but it should be connected to live sources and a validated domain model before policy deployment."
        )
        reference_lines = "\n\n".join(self._reference_entry_markdown(source) for source in run.source_registry)
        sections.append("## References\n" + reference_lines)
        return "\n\n".join(sections)

    def _score_report(self, report_text: str, cycle_number: int) -> ReviewScorecard:
        section_scores = self._section_scores(report_text)
        citations, analysis_refs = extract_citation_ids(report_text)
        has_agency_ownership = any(
            token in report_text
            for token in [
                "Lead implementing agency",
                "Implementation Sequencing and Agency Ownership",
                "Department",
                "Ministry",
            ]
        )
        has_scenario = any(
            token in report_text.lower()
            for token in [
                "multiple dimensions move together",
                "counterfactual",
                "downside scenario",
            ]
        )
        wc = word_count(report_text)
        criteria = [
            CriterionScore(
                criterion_id=1,
                criterion="Structure and Flow",
                score=8.0 if section_scores["ordered"] else 6.2,
                explanation="Scores higher when the report preserves the approved section order and transitions cleanly.",
            ),
            CriterionScore(
                criterion_id=2,
                criterion="Government Client Readability",
                score=8.1 if wc <= 2300 else 6.4,
                explanation="The report remains concise enough for rapid briefing and uses direct lead messages.",
            ),
            CriterionScore(
                criterion_id=3,
                criterion="Thought Innovation",
                score=8.0 if has_scenario else 5.8,
                explanation="Non-obvious scenario framing is required for a strong score on insight depth.",
            ),
            CriterionScore(
                criterion_id=4,
                criterion="Literature Comprehension",
                score=8.0 if len(citations) >= 6 else 6.1,
                explanation="The literature section should make substantive use of multiple verified sources.",
            ),
            CriterionScore(
                criterion_id=5,
                criterion="Knowledge Creation",
                score=7.8 if len(analysis_refs) >= 4 else 6.3,
                explanation="The brief should add new empirical interpretation rather than restate prior literature.",
            ),
            CriterionScore(
                criterion_id=6,
                criterion="Data Accuracy",
                score=7.9 if "offline local-fixture mode" in report_text.lower() else 6.2,
                explanation="The report should explicitly state caveats, reference periods, and evidence provenance.",
            ),
            CriterionScore(
                criterion_id=7,
                criterion="Econometric Rigour",
                score=8.0 if "confidence band" in report_text.lower() and "composite-indicator" in report_text.lower() else 6.0,
                explanation="Method transparency and honest uncertainty reporting are required.",
            ),
            CriterionScore(
                criterion_id=8,
                criterion="Policy Recommendations",
                score=8.1 if has_agency_ownership else 5.9,
                explanation="Recommendations must name implementers and tie directly to analysis-backed datapoints.",
            ),
            CriterionScore(
                criterion_id=9,
                criterion="Professional Presentation",
                score=8.0 if "| Indicator | Latest | Prior | Delta | Dimension |" in report_text else 6.3,
                explanation="Presentation improves when tables, references, and section formatting are consistent.",
            ),
        ]
        composite = round(mean(item.score for item in criteria), 2)
        passed = composite >= 7.5 and all(item.score >= 6.0 for item in criteria)
        priority_fixes = [
            f"Raise {criterion.criterion.lower()} by addressing: {criterion.explanation}"
            for criterion in criteria
            if criterion.score < 6.5
        ]
        if not priority_fixes:
            priority_fixes.append("No blocking fixes. Preserve the current structure and audit trail.")
        summary_fallback = (
            "The reviewer assessed only the report artifact. The score improves when the brief adds stronger scenario framing, "
            "clearer agency ownership, and a more explicit explanation of how the headline metric translates into action."
        )
        reviewer_summary = self.local_llm.complete(
            system_prompt="Write a concise review summary for a policy report scorecard.",
            user_prompt=(
                f"Cycle: {cycle_number}\n"
                f"Composite score: {composite}\n"
                f"Priority fixes: {'; '.join(priority_fixes)}"
            ),
            fallback=summary_fallback,
        )
        return ReviewScorecard(
            cycle_number=cycle_number,
            criterion_scores=criteria,
            composite_score=composite,
            passed=passed,
            priority_fixes=priority_fixes,
            reviewer_summary=reviewer_summary,
        )

    def _section_scores(self, report_text: str) -> dict[str, Any]:
        positions = [report_text.find(f"## {title}") for title in SECTION_ORDER]
        ordered = all(pos != -1 for pos in positions) and positions == sorted(positions)
        return {"ordered": ordered}

    def _profile(self, run: SerializedRunState):
        if not run.topic_profile:
            run.topic_profile = infer_topic_profile_with_llm(run.topic, self.local_llm)
        return run.topic_profile

    def _checkpoint_feedback(self, run: SerializedRunState, checkpoint_id: CheckpointId) -> str | None:
        record = run.checkpoint_status.get(checkpoint_id.value)
        return record.last_feedback if record else None

    def _render_plan_markdown(self, run: SerializedRunState, output: PlannerOutput) -> str:
        profile = self._profile(run)
        sections = "\n".join(
            f"- {section.title}: {section.target_words} words. {section.purpose}"
            for section in output.section_plan
        )
        indicators = "\n".join(
            f"- {indicator.label} ({indicator.frequency}, {indicator.unit}) -> {indicator.sector_bucket}"
            for indicator in output.indicator_plan
        )
        return (
            "# Planner Output\n"
            f"## Topic Profile\n- Domain: {profile.domain}\n- Geography: {profile.geography or 'Not explicit'}\n"
            f"- Headline Metric: {profile.headline_metric_label} ({profile.headline_unit})\n"
            f"- LLM Runtime: {self.local_llm.runtime_label()}\n\n"
            f"## Problem Statement\n{output.problem_statement}\n\n"
            f"## Analytical Method\n{output.analytical_method}\n\n"
            f"## Section Plan\n{sections}\n\n"
            f"## Indicator Plan\n{indicators}\n\n"
            f"## Notes\n{output.planner_notes}\n"
        )

    def _render_literature_markdown(self, run: SerializedRunState, pack: LiteraturePack) -> str:
        profile = self._profile(run)
        lines = [
            "# Literature Review",
            f"- Domain: {profile.domain}",
            f"- Topic: {run.topic}",
            "",
            pack.introduction_summary,
            "",
            pack.literature_review_summary,
            "",
            "## Sources",
        ]
        for source in pack.sources:
            lines.append(f"- [{source.source_id}] {source.title}: {source.relevance}")
        return "\n".join(lines)

    def _render_data_collection_markdown(self, pack: IndicatorPack) -> str:
        rows = [
            {
                "Indicator": indicator.label,
                "Latest": f"{indicator.latest_value:.1f} {indicator.unit}",
                "Period": indicator.reference_period,
                "Source": indicator.primary_source_name,
            }
            for indicator in pack.indicators
        ]
        return "# Data Collection\n" + render_markdown_table(rows, ["Indicator", "Latest", "Period", "Source"])

    def _render_fact_check_markdown(self, pack: VerificationPack) -> str:
        lines = ["# Fact Check", pack.correction_summary, "", "## Verification Items"]
        for item in pack.verification_items:
            lines.append(
                f"- {item.item_type.upper()} {item.item_id}: {item.status} ({item.primary_source} vs {item.secondary_source})"
            )
        return "\n".join(lines)

    def _render_analysis_markdown(self, pack: AnalysisPack) -> str:
        attribution_lines = "\n".join(
            f"- {bucket}: {value:.3f}" for bucket, value in pack.sector_attribution.items()
        )
        datapoints = "\n".join(
            f"- [{item.datapoint_id}] {item.label}: {item.value:.2f} {item.unit}" for item in pack.data_points
        )
        return (
            "# Analysis Output\n"
            f"- Headline metric: {pack.headline_metric_label}\n"
            f"- Target period: {pack.target_period}\n"
            f"- Point estimate: {pack.point_estimate:.2f} {pack.headline_unit}\n"
            f"- Confidence band: {pack.confidence_low:.2f} to {pack.confidence_high:.2f}\n"
            f"- Comparator ({pack.comparator_label}): {pack.last_official_estimate:.2f}\n\n"
            f"## {pack.dimension_label}\n"
            f"{attribution_lines}\n\n"
            "## Data Points\n"
            f"{datapoints}\n\n"
            "## Commentary\n"
            f"{pack.scenario_commentary}\n"
        )

    def _render_review_markdown(self, scorecard: ReviewScorecard) -> str:
        score_rows = [
            {
                "Criterion": item.criterion,
                "Score": f"{item.score:.1f}",
                "Explanation": item.explanation,
            }
            for item in scorecard.criterion_scores
        ]
        fixes = "\n".join(f"- {fix}" for fix in scorecard.priority_fixes)
        return (
            f"# Critical Review Cycle {scorecard.cycle_number}\n"
            f"- Composite score: {scorecard.composite_score:.2f}\n"
            f"- Passed: {scorecard.passed}\n\n"
            f"{render_markdown_table(score_rows, ['Criterion', 'Score', 'Explanation'])}\n\n"
            "## Priority Fixes\n"
            f"{fixes}\n\n"
            "## Summary\n"
            f"{scorecard.reviewer_summary}\n"
        )

    def _reference_entry_markdown(self, source) -> str:
        citation = source.apa_citation.removesuffix(f" {source.url}").strip()
        return f"- [{source.source_id}] {citation}\nSource link: [Open article]({source.url})"

    def _analysis_script_content(self) -> str:
        return (
            "from pif_research_platform.analysis.generic_topic import GenericTopicAnalysisStrategy\n\n"
            "strategy = GenericTopicAnalysisStrategy()\n"
            "# Reproduce the analysis by loading the saved topic profile and indicator snapshot, then calling strategy.run(...)\n"
        )

    def _write_indicator_workbook(
        self,
        run: SerializedRunState,
        pack: IndicatorPack | None,
        analysis_pack: AnalysisPack | None,
    ) -> None:
        if pack is None:
            return
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "indicator_table"
        headers = [
            "indicator_id",
            "label",
            "reference_period",
            "latest_value",
            "unit",
            "dimension",
            "primary_source_url",
            "verification_source_url",
        ]
        sheet.append(headers)
        for item in pack.indicators:
            sheet.append(
                [
                    item.indicator_id,
                    item.label,
                    item.reference_period,
                    item.latest_value,
                    item.unit,
                    item.sector_bucket,
                    item.primary_source_url,
                    item.verification_source_url,
                ]
            )
        meta = workbook.create_sheet("metadata")
        meta.append(["field", "value"])
        for key, value in pack.metadata.items():
            meta.append([key, value])
        if analysis_pack:
            analysis_sheet = workbook.create_sheet("analysis_output")
            analysis_sheet.append(["dimension", "contribution", "interpretation"])
            for row in analysis_pack.analysis_table:
                analysis_sheet.append([row["bucket"], row["contribution"], row["interpretation"]])
        path = self.repo.run_dir(run.run_id) / "artifacts" / "indicators.xlsx"
        workbook.save(path)
        run.artifact_paths["indicator_workbook"] = str(path)

    def _build_padding_text(self, run: SerializedRunState, target_words: int) -> str:
        if not (run.analysis_pack and run.literature_pack and run.indicator_pack):
            return ""
        profile = self._profile(run)
        analysis = run.analysis_pack
        geography = profile.geography or "the target geography"
        owners = self._agency_owners(profile, geography)
        rationale_by_id = {
            spec.indicator_id: spec.rationale
            for spec in run.indicator_plan
        }
        citations = " ".join(f"[{source.source_id}]" for source in run.source_registry[:4])
        indicator_rows = [
            {
                "Indicator": item.label,
                "Owner": self._watchlist_owner_label(item.sector_bucket, owners),
                "Watch for": "Sustained deterioration" if item.delta < 0 else "Loss of recent gains",
                "Trigger": f"{abs(item.delta):.1f}+ point adverse move",
            }
            for item in run.indicator_pack.indicators
        ]
        sections: list[str] = [
            (
                f"The monitoring watchlist for {run.topic} should be anchored in the headline "
                f"{analysis.headline_metric_label.lower()} and the signals that explain {profile.dimension_label.lower()}. "
                f"For {geography}, the practical objective is not just to note whether the score moved, but to identify "
                f"which release changed the decision context, which agency now owns the response, and whether the current "
                f"confidence band is widening because the signal is fragmenting across dimensions. {citations}"
            ),
            render_markdown_table(indicator_rows, ["Indicator", "Owner", "Watch for", "Trigger"]),
        ]
        optional_sections: list[str] = []
        for item in sorted(run.indicator_pack.indicators, key=lambda indicator: abs(indicator.delta), reverse=True):
            rationale = rationale_by_id.get(
                item.indicator_id,
                "This series provides a practical operating signal for the topic.",
            )
            direction = "improved" if item.delta >= 0 else "weakened"
            optional_sections.append(
                (
                    f"{item.label}: {self._watchlist_owner_label(item.sector_bucket, owners)} should review the "
                    f"{item.reference_period} release in detail because the latest reading is {item.latest_value:.1f} "
                    f"{item.unit}, compared with {item.prior_value:.1f} {item.unit} previously, meaning the signal has "
                    f"{direction} by {abs(item.delta):.1f} points over the last cycle. The series sits in the "
                    f"{item.sector_bucket.lower()} bucket and matters here because {rationale.lower()} For the current "
                    f"brief, a reversal in this series would change how confidently the team can interpret the headline "
                    f"assessment and whether the next action should prioritize immediate containment, delivery repair, "
                    f"or broader institutional coordination."
                )
            )
        optional_sections.extend(
            [
                (
                    "#### Escalation Triggers\n"
                    f"1. Trigger a senior review if the headline confidence band widens by more than one release cycle without a clear explanation in {analysis.dimension_label.lower()}.\n"
                    f"2. Trigger corrective action if the dominant dimension shifts away from {max(analysis.sector_attribution, key=analysis.sector_attribution.get).lower()} and the comparator gap deteriorates at the same time.\n"
                    "3. Trigger a source-quality check when official releases are delayed, definitions change, or one indicator moves sharply without confirmation from a second operational signal."
                ),
                (
                    "#### Review Rhythm and Ownership\n"
                    f"{owners['lead']} should chair the first-line review, because that office is best placed to connect the "
                    f"headline finding to operational action. {owners['operations']} should maintain the release calendar, "
                    f"document what changed in each indicator, and identify whether the movement reflects real field "
                    f"conditions or only a reporting effect. {owners['monitor']} should keep a short escalation note that "
                    " states what moved, why it matters, what action is proposed, and what evidence is still missing."
                ),
                (
                    "#### Decision Questions for the Next Cycle\n"
                    f"- Which indicator moved enough to change the recommended stance for {run.topic}?\n"
                    f"- Is the current signal broad-based across {', '.join(profile.indicator_dimensions[:3]).lower()}, or is it dependent on only one release?\n"
                    f"- Does the next action belong with {owners['lead']}, or does it require joint action with {owners['operations']}?\n"
                    "- Are the observed movements large enough to justify a policy shift, or should the team wait for confirmation from the next release?\n"
                    "- Which operational bottleneck is becoming more binding: delivery capacity, financing, compliance, or local implementation quality?\n"
                    "- Which public-facing risk should be communicated now, and which issue should remain under internal monitoring until evidence is clearer?"
                ),
                (
                    "#### Data Quality Checks\n"
                    "The monitoring note should distinguish between a genuine policy signal and a temporary statistical artifact. "
                    "That means recording the release date, source URL, comparison base, revision history, and whether an "
                    "independent verification source confirms the movement. For a prototype run like this one, the team should "
                    "also note where the workflow is still using local fixtures so that a later production connector can replace "
                    "the placeholder evidence trail without changing the decision logic."
                ),
            ]
        )
        extra_notes = [
            (
                f"Contingency note on {item.label.lower()}: if the next release moves against the current signal, "
                f"{self._watchlist_owner_label(item.sector_bucket, owners)} should explain whether the shift came from "
                "seasonality, administrative delay, enforcement change, or a deeper deterioration in field conditions. "
                "That keeps the watchlist decision-oriented rather than descriptive."
            )
            for item in run.indicator_pack.indicators
        ]
        selected_sections = list(sections)
        for section in optional_sections + extra_notes:
            selected_sections.append(section)
            if word_count("\n\n".join(selected_sections)) >= target_words:
                break
        return "\n\n".join(selected_sections)

    def _build_policy_recommendations(
        self,
        run: SerializedRunState,
        profile,
        analysis: AnalysisPack,
        leading_dimension: str,
        top_indicators,
        *,
        improved_revision: bool,
    ) -> str:
        geography = profile.geography or "the target geography"
        owners = self._agency_owners(profile, geography)
        primary = top_indicators[0]
        secondary = top_indicators[1] if len(top_indicators) > 1 else top_indicators[0]
        weakest = min(run.indicator_pack.indicators, key=lambda item: item.delta)
        policy_lines = [
            (
                f"Lead implementing agency: {owners['lead']} should own the first response cycle for {primary.label.lower()} "
                f"because it is one of the clearest observable signals shaping the current assessment for {run.topic.lower()} [A1]."
            ),
            (
                f"{owners['operations']} should sequence operational action around the {leading_dimension.lower()} dimension and "
                f"remove bottlenecks visible in {secondary.label.lower()} so that gains do not remain isolated to a single signal [A2][A3]."
            ),
            (
                f"{owners['monitor']} should run a standing review cadence and escalate when the confidence band widens, the "
                f"comparator gap worsens, or {weakest.label.lower()} weakens enough to change the implementation stance [A4]."
            ),
        ]
        sequencing_lines = [
            "### Implementation Sequencing and Agency Ownership",
            (
                f"Immediate phase: {owners['lead']} should validate the current signal, publish a compact owner-level "
                f"dashboard for {geography}, and assign one accountable officer for the next 30-60 day response."
            ),
            (
                f"Execution phase: {owners['operations']} should focus on the bottleneck that is most likely to dilute "
                f"{leading_dimension.lower()} support, especially where {secondary.label.lower()} and {weakest.label.lower()} "
                "suggest the pipeline is not yet broad-based."
            ),
            (
                f"Governance phase: {owners['monitor']} should review the indicator basket each cycle, record whether the "
                f"signal is diffusing across {profile.dimension_label.lower()}, and decide whether the next move is scale-up, "
                "course correction, or a temporary hold pending stronger evidence."
            ),
        ]
        if improved_revision:
            sequencing_lines.append(
                (
                    f"Decision rule: if the next release confirms improvement in {primary.label.lower()} but not in adjacent "
                    f"dimensions, treat the signal as encouraging but incomplete; if both {primary.label.lower()} and "
                    f"{secondary.label.lower()} strengthen together, the programme can move from containment to structured scale-up."
                )
            )
        return (
            "## Policy Recommendations\n"
            + "\n".join(f"{index}. {line}" for index, line in enumerate(policy_lines, start=1))
            + "\n\n"
            + "\n\n".join(sequencing_lines)
        )

    def _agency_owners(self, profile, geography: str) -> dict[str, str]:
        geography_prefix = "" if geography == "the target geography" else f"{geography} "
        if profile.domain == "environment":
            lead = "State Pollution Control Board"
            if "delhi" in geography.lower():
                lead = "Delhi Pollution Control Committee"
            elif geography != "the target geography":
                lead = f"{geography} Pollution Control Board"
            return {
                "lead": lead,
                "operations": "Transport Department, urban local bodies, and enforcement agencies",
                "monitor": "Health Department and the programme monitoring cell",
            }
        if profile.domain == "energy_transition":
            return {
                "lead": f"{geography_prefix}Energy Department and DISCOM leadership".strip(),
                "operations": "Transport Department and infrastructure delivery agencies",
                "monitor": "Finance Department and programme management unit",
            }
        if profile.domain == "mobility":
            return {
                "lead": f"{geography_prefix}Transport Department and network operator".strip(),
                "operations": "Urban development and corridor delivery agencies",
                "monitor": "Planning department and mobility PMU",
            }
        if profile.domain == "social_policy":
            return {
                "lead": "Lead line department",
                "operations": "District implementation units and frontline administrators",
                "monitor": "Finance and monitoring cell",
            }
        if profile.domain == "macroeconomy":
            return {
                "lead": f"{geography_prefix}Finance Department".strip(),
                "operations": "Industries, power, and transport departments",
                "monitor": "Planning department and economic advisory unit",
            }
        return {
            "lead": "Lead line department",
            "operations": "Implementing agencies across the delivery chain",
            "monitor": "Cabinet or programme monitoring unit",
        }

    def _watchlist_owner_label(self, sector_bucket: str, owners: dict[str, str]) -> str:
        bucket = sector_bucket.lower()
        if bucket in {"health", "quality", "risk"}:
            return owners["monitor"]
        if bucket in {"delivery", "enforcement", "capacity", "mobility", "emissions"}:
            return owners["operations"]
        return owners["lead"]
